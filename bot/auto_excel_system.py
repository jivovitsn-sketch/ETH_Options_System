#!/usr/bin/env python3
"""
AUTO EXCEL SYSTEM
- Автозаполнение по сигналу
- Редактирование начальных данных
- Ручной ввод TP и профита
- Синхронизация с Discord
"""
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from discord_multi_channel import DiscordMultiChannel

class AutoExcelSystem:
    def __init__(self):
        self.excel_file = Path(__file__).parent.parent / 'data' / 'Live_Trading_Journal.xlsx'
        self.discord = DiscordMultiChannel()
        
        # Создаём Excel если нет
        if not self.excel_file.exists():
            self._create_excel_template()
        
        print("✅ Auto Excel System initialized")
        print(f"   Excel: {self.excel_file}")
    
    def _create_excel_template(self):
        """Создать шаблон Excel"""
        
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Live Trades"
        
        # Заголовок
        headers = [
            '№', 'Статус', 'Дата Сигнала', 'Актив', 'Стратегия', 'Направление',
            'Цена Входа', 'Кол-во/Размер', 'Strike', 'DTE', 'Expiration',
            'Delta', 'Theta', 'Gamma', 'Vega', 'IV (%)',
            'TP1 (цена)', 'TP1 ($)', 'TP2 (цена)', 'TP2 ($)', 'TP3 (цена)', 'TP3 ($)',
            'Цена Закр.', 'Дата Закр.', 'Факт. Профит ($)', 'Факт. Профит (%)',
            'Комиссия', 'Чистый Профит', 'Hold Time (h)', 'Причина Выхода',
            'Капитал', 'Заметки'
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # Ширина колонок
        widths = {
            'A': 5, 'B': 10, 'C': 12, 'D': 12, 'E': 15, 'F': 10,
            'G': 12, 'H': 12, 'I': 10, 'J': 8, 'K': 12,
            'L': 8, 'M': 8, 'N': 8, 'O': 8, 'P': 8,
            'Q': 12, 'R': 12, 'S': 12, 'T': 12, 'U': 12, 'V': 12,
            'W': 12, 'X': 12, 'Y': 15, 'Z': 15,
            'AA': 10, 'AB': 15, 'AC': 12, 'AD': 15,
            'AE': 12, 'AF': 30
        }
        
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        
        wb.save(self.excel_file)
        print(f"✅ Created Excel template: {self.excel_file}")
    
    def add_signal(self, signal_data):
        """
        Добавить новый сигнал (АВТОЗАПОЛНЕНИЕ)
        
        signal_data = {
            'asset': 'ETHUSDT',
            'strategy': 'Bull Call 60DTE',
            'direction': 'LONG',
            'entry_price': 3939.00,
            'size': 500,
            'strike': 4000,
            'dte': 60,
            'expiration': '2025-12-31',
            'delta': 0.071,
            'theta': 0.05,
            'gamma': 0.001,
            'vega': 0.15,
            'iv': 65.5
        }
        """
        
        wb = load_workbook(self.excel_file)
        ws = wb.active
        
        # Найти следующую пустую строку
        next_row = ws.max_row + 1
        trade_id = next_row - 1
        
        # Автозаполнение (пользователь может редактировать)
        ws.cell(row=next_row, column=1, value=trade_id)  # №
        ws.cell(row=next_row, column=2, value='OPEN')  # Статус
        ws.cell(row=next_row, column=2).font = Font(color='FFA500', bold=True)
        
        ws.cell(row=next_row, column=3, value=datetime.now().strftime('%Y-%m-%d %H:%M'))  # Дата
        ws.cell(row=next_row, column=4, value=signal_data['asset'])  # Актив
        ws.cell(row=next_row, column=5, value=signal_data['strategy'])  # Стратегия
        ws.cell(row=next_row, column=6, value=signal_data.get('direction', 'LONG'))  # Направление
        ws.cell(row=next_row, column=7, value=signal_data['entry_price'])  # Цена входа
        ws.cell(row=next_row, column=8, value=signal_data['size'])  # Размер
        ws.cell(row=next_row, column=9, value=signal_data.get('strike', ''))  # Strike
        ws.cell(row=next_row, column=10, value=signal_data.get('dte', ''))  # DTE
        ws.cell(row=next_row, column=11, value=signal_data.get('expiration', ''))  # Expiration
        ws.cell(row=next_row, column=12, value=signal_data.get('delta', ''))  # Delta
        ws.cell(row=next_row, column=13, value=signal_data.get('theta', ''))  # Theta
        ws.cell(row=next_row, column=14, value=signal_data.get('gamma', ''))  # Gamma
        ws.cell(row=next_row, column=15, value=signal_data.get('vega', ''))  # Vega
        ws.cell(row=next_row, column=16, value=signal_data.get('iv', ''))  # IV
        
        # TP уровни - ПУСТЫЕ (заполняет пользователь вручную!)
        # Колонки Q-V остаются пустыми
        
        # Подсветка ячеек для ручного ввода
        for col in range(17, 23):  # Q-V (TP levels)
            cell = ws.cell(row=next_row, column=col)
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        wb.save(self.excel_file)
        
        print(f"✅ Signal added to Excel: Trade #{trade_id}")
        
        # Отправка в Discord
        self.discord.vip_trade_opened({
            'asset': signal_data['asset'],
            'strategy': signal_data['strategy'],
            'entry': signal_data['entry_price'],
            'size': signal_data['size'],
            'sl': 0,
            'tp': 0,
            'rr': 2,
            'delta': signal_data.get('delta', 0),
            'theta': signal_data.get('theta', 0),
            'quality': 'High'
        })
        
        self.discord.free_signal(
            signal_data['asset'],
            'BUY' if signal_data.get('direction') == 'LONG' else 'SELL',
            signal_data['entry_price']
        )
        
        return trade_id
    
    def close_trade(self, trade_id, exit_data):
        """
        Закрыть сделку (ФИНАЛЬНЫЕ ДАННЫЕ)
        
        exit_data = {
            'exit_price': 4050.00,
            'actual_profit': 55.50,
            'actual_profit_pct': 11.1,
            'commission': 0.60,
            'exit_reason': 'TP2',
            'notes': 'Good trade'
        }
        """
        
        wb = load_workbook(self.excel_file)
        ws = wb.active
        
        # Найти строку по trade_id
        trade_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == trade_id:
                trade_row = row
                break
        
        if not trade_row:
            print(f"❌ Trade #{trade_id} not found")
            return
        
        # Обновление данных
        ws.cell(row=trade_row, column=2, value='CLOSED')  # Статус
        ws.cell(row=trade_row, column=2).font = Font(color='00FF00' if exit_data['actual_profit'] > 0 else 'FF0000', bold=True)
        
        ws.cell(row=trade_row, column=23, value=exit_data['exit_price'])  # Цена закрытия
        ws.cell(row=trade_row, column=24, value=datetime.now().strftime('%Y-%m-%d %H:%M'))  # Дата закрытия
        ws.cell(row=trade_row, column=25, value=exit_data['actual_profit'])  # Факт. профит
        ws.cell(row=trade_row, column=26, value=exit_data['actual_profit_pct'] / 100)  # Факт. профит %
        ws.cell(row=trade_row, column=27, value=exit_data.get('commission', 0))  # Комиссия
        
        net_profit = exit_data['actual_profit'] - exit_data.get('commission', 0)
        ws.cell(row=trade_row, column=28, value=net_profit)  # Чистый профит
        
        # Hold time
        date_open = ws.cell(row=trade_row, column=3).value
        if isinstance(date_open, str):
            date_open = datetime.strptime(date_open, '%Y-%m-%d %H:%M')
        hold_hours = (datetime.now() - date_open).total_seconds() / 3600
        ws.cell(row=trade_row, column=29, value=hold_hours)
        
        ws.cell(row=trade_row, column=30, value=exit_data.get('exit_reason', ''))  # Причина
        ws.cell(row=trade_row, column=32, value=exit_data.get('notes', ''))  # Заметки
        
        # Цвет для профита
        profit_cell = ws.cell(row=trade_row, column=25)
        profit_cell.font = Font(color='00FF00' if exit_data['actual_profit'] > 0 else 'FF0000', bold=True)
        
        wb.save(self.excel_file)
        
        print(f"✅ Trade #{trade_id} closed in Excel")
        
        # Discord уведомление
        entry_price = ws.cell(row=trade_row, column=7).value
        asset = ws.cell(row=trade_row, column=4).value
        
        self.discord.vip_trade_closed({
            'asset': asset,
            'entry': entry_price,
            'exit': exit_data['exit_price'],
            'pnl': exit_data['actual_profit'],
            'pnl_pct': exit_data['actual_profit_pct'],
            'hold_hours': hold_hours,
            'exit_reason': exit_data.get('exit_reason', ''),
            'commission': exit_data.get('commission', 0),
            'capital': 10000 + exit_data['actual_profit']  # Simplified
        })
    
    def sync_with_journal(self):
        """Синхронизировать Excel с JSON журналом"""
        
        wb = load_workbook(self.excel_file)
        ws = wb.active
        
        trades = []
        
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row=row, column=2).value
            
            if status == 'CLOSED':
                trade = {
                    'id': ws.cell(row=row, column=1).value,
                    'timestamp': str(ws.cell(row=row, column=3).value),
                    'asset': ws.cell(row=row, column=4).value,
                    'strategy': ws.cell(row=row, column=5).value,
                    'entry_price': ws.cell(row=row, column=7).value,
                    'exit_price': ws.cell(row=row, column=23).value,
                    'size': ws.cell(row=row, column=8).value,
                    'pnl': ws.cell(row=row, column=25).value,
                    'pnl_pct': ws.cell(row=row, column=26).value * 100 if ws.cell(row=row, column=26).value else 0,
                    'commission': ws.cell(row=row, column=27).value or 0,
                    'hold_time_hours': ws.cell(row=row, column=29).value or 0,
                    'exit_reason': ws.cell(row=row, column=30).value or '',
                    'capital_after': 10000,  # Calculate properly
                    'notes': ws.cell(row=row, column=32).value or ''
                }
                
                trades.append(trade)
        
        # Save to JSON
        import json
        journal_file = Path(__file__).parent.parent / 'data' / 'trade_journal.json'
        
        with open(journal_file, 'w') as f:
            json.dump(trades, f, indent=2)
        
        print(f"✅ Synced {len(trades)} closed trades to journal")
        
        return trades

# DEMO
if __name__ == "__main__":
    system = AutoExcelSystem()
    
    print("\n" + "="*80)
    print("DEMO: Автозаполнение по сигналу")
    print("="*80)
    
    # 1. Новый сигнал (автозаполнение)
    signal = {
        'asset': 'ETHUSDT',
        'strategy': 'Bull Call 60DTE',
        'direction': 'LONG',
        'entry_price': 3939.00,
        'size': 500,
        'strike': 4000,
        'dte': 60,
        'expiration': '2025-12-31',
        'delta': 0.071,
        'theta': 0.05,
        'gamma': 0.001,
        'vega': 0.15,
        'iv': 65.5
    }
    
    trade_id = system.add_signal(signal)
    
    print(f"\n✅ Добавлен сигнал #{trade_id}")
    print(f"   Открой Excel: {system.excel_file}")
    print(f"   Заполни вручную:")
    print(f"     - TP1, TP2, TP3 (цены и суммы)")
    print(f"     - Можешь редактировать начальные данные")
    
    print(f"\n📊 Excel файл: {system.excel_file}")
